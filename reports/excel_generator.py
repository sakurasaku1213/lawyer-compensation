#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel・帳票出力エンジン（最適化版）
プロフェッショナルな書類作成機能

改善点:
- 設定システムとの統合
- エラーハンドリングの強化
- パフォーマンス最適化
- テンプレート機能の充実
- データ検証の強化
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.drawing.image import Image
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, Any, List, Optional, Union
import os
import logging
import json
from pathlib import Path
from dataclasses import asdict

from models import CaseData
from calculation.compensation_engine import CalculationResult
from config.app_config import AppConfig, ReportConfig
from utils import (
    ErrorHandler, get_error_handler, CompensationSystemError, 
    ErrorCategory, ErrorSeverity, FileIOError, ConfigurationError,
    monitor_performance, get_performance_monitor
)


class ExcelReportGeneratorOptimized:
    """Excel帳票生成クラス（最適化版）"""

    def __init__(self, config: Optional[AppConfig] = None):
        """初期化"""
        self.config = config or AppConfig()
        self.report_config = self.config.report
        self.logger = logging.getLogger(__name__)
        self.error_handler = get_error_handler()
        self.performance_monitor = get_performance_monitor()
        
        # 出力ディレクトリの設定と作成
        self._setup_directories()
        
        # スタイル設定の初期化（設定ファイルから読み込み）
        self._initialize_styles()
        
        # テンプレートマネージャーの初期化
        self.template_manager = TemplateManager(self.report_config, self.error_handler)
        
        # 自動テンプレート作成設定
        if self.report_config.auto_create_missing_templates:
            self.template_manager.create_standard_templates()
        
        self.logger.info("Excel帳票生成システム（最適化版）を初期化しました")

    def _setup_directories(self):
        """ディレクトリの設定と作成"""
        try:
            # 出力ディレクトリ
            self.output_dir = Path(self.report_config.output_directory)
            self.output_dir.mkdir(parents=True, exist_ok=True)
            
            # テンプレートディレクトリ
            if self.report_config.excel_templates:
                default_template_path = self.report_config.excel_templates.get("default")
                if default_template_path:
                    self.template_dir = Path(default_template_path).parent
                else:
                    self.template_dir = Path("templates/excel")
            else:
                if self.report_config.excel_template_path:
                    self.template_dir = Path(self.report_config.excel_template_path).parent
                else:
                    self.template_dir = Path("templates/excel")
            
            self.template_dir.mkdir(parents=True, exist_ok=True)
            
        except OSError as e:
            self.error_handler.handle_exception(
                FileIOError(f"ディレクトリの作成に失敗しました: {e}",
                            user_message="レポート用ディレクトリの準備中にエラーが発生しました。",
                            context={"output_dir": str(self.output_dir), "template_dir": str(self.template_dir)})
            )
            # フォールバック
            self.output_dir = Path("./reports")
            self.template_dir = Path("./templates")
            self.output_dir.mkdir(parents=True, exist_ok=True)
            self.template_dir.mkdir(parents=True, exist_ok=True)

    def _initialize_styles(self):
        """Excelスタイル設定の初期化（設定ファイルから読み込み）"""
        try:
            colors = self.report_config.excel_color_scheme
            default_font = self.config.ui.font_family
            
            # フォント設定
            self.fonts = {
                'title': Font(name=default_font, size=18, bold=True, color=colors.get('header_text', 'FFFFFF')),
                'header': Font(name=default_font, size=14, bold=True, color=colors.get('header_text', 'FFFFFF')),
                'subtitle': Font(name=default_font, size=12, bold=True, color=colors.get('body_text', '000000')),
                'body': Font(name=default_font, size=10, color=colors.get('body_text', '000000')),
                'number': Font(name=default_font, size=10, color=colors.get('body_text', '000000')),
                'small': Font(name=default_font, size=8, color=colors.get('body_text', '000000')),
                'money': Font(name=default_font, size=12, bold=True, color='C55A5A'),
                'important': Font(name=default_font, size=11, bold=True, color='E74C3C')
            }
            
            # 塗りつぶし設定
            self.fills = {
                'header': PatternFill(start_color=colors.get('header_bg', '4472C4'), 
                                    end_color=colors.get('header_bg', '4472C4'), 
                                    fill_type='solid'),
                'subheader': PatternFill(start_color=colors.get('subheader_bg', '8DB4E2'), 
                                       end_color=colors.get('subheader_bg', '8DB4E2'), 
                                       fill_type='solid'),
                'amount': PatternFill(start_color=colors.get('amount_bg', 'FFF2CC'), 
                                    end_color=colors.get('amount_bg', 'FFF2CC'), 
                                    fill_type='solid'),
                'total': PatternFill(start_color=colors.get('total_bg', 'D5E8D4'), 
                                   end_color=colors.get('total_bg', 'D5E8D4'), 
                                   fill_type='solid')
            }
            
            # 罫線設定
            self.borders = {
                'thin_all': Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin')),
                'medium_all': Border(left=Side(style='medium'), right=Side(style='medium'),
                                   top=Side(style='medium'), bottom=Side(style='medium')),
                'thick_bottom': Border(bottom=Side(style='thick'))
            }
            
            # 配置設定
            self.alignments = {
                'left': Alignment(horizontal='left', vertical='center'),
                'center': Alignment(horizontal='center', vertical='center'),
                'right': Alignment(horizontal='right', vertical='center'),
                'center_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True)
            }
            
            # 列幅と行高設定
            self.column_widths = self.report_config.excel_column_widths
            self.row_heights = self.report_config.excel_row_heights
            
            self.logger.debug("Excelスタイル設定を初期化しました")
        except Exception as e:
            self.error_handler.handle_exception(
                ConfigurationError(f"Excelスタイル設定の初期化に失敗しました: {e}",
                                   user_message="Excelレポートのスタイル設定の読み込みに失敗しました。",
                                   context={"error_details": str(e)})
            )
            self._set_default_styles()

    def _set_default_styles(self):
        """デフォルトスタイルの設定（フォールバック用）"""
        default_font = self.config.ui.font_family
        
        self.fonts = {
            'title': Font(name=default_font, size=16, bold=True, color='1F4E79'),
            'subtitle': Font(name=default_font, size=14, bold=True, color='2E75B6'),
            'header': Font(name=default_font, size=12, bold=True, color='FFFFFF'),
            'body': Font(name=default_font, size=11),
            'small': Font(name=default_font, size=10, color='666666'),
            'number': Font(name=default_font, size=11, bold=True),
            'money': Font(name=default_font, size=12, bold=True, color='C55A5A'),
            'important': Font(name=default_font, size=11, bold=True, color='E74C3C')
        }
        
        self.borders = {
            'thin_all': Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin')),
            'thick_bottom': Border(bottom=Side(style='thick')),
            'medium_all': Border(left=Side(style='medium'), right=Side(style='medium'),
                               top=Side(style='medium'), bottom=Side(style='medium'))
        }
        
        self.fills = {
            'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'subheader': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
            'total': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'amount': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        }
        
        self.alignments = {
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center'),
            'center_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True)
        }
        
        self.column_widths = {"A": 20.0, "B": 15.0, "C": 40.0, "D": 25.0}
        self.row_heights = {"header": 25.0, "subheader": 20.0, "data": 18.0, "footer": 15.0}

    @monitor_performance("excel_report_generation", track_parameters=True)
    def create_compensation_report(self, case_data: CaseData, results: Dict[str, CalculationResult], 
                                 output_filename: str, template_type: str = "default") -> bool:
        """損害賠償計算書の作成 - パフォーマンス最適化版"""
        try:
            self.logger.info(f"損害賠償計算書の作成を開始します: {output_filename}")
            
            # テンプレートが指定されている場合は適用
            wb = None
            if self.report_config.enable_template_customization and template_type != "none":
                wb = self.template_manager.apply_template(template_type, case_data)
                if wb is None:
                    self.logger.warning(f"テンプレート '{template_type}' の適用に失敗したため、新規作成します")
            
            if wb is None:
                wb = openpyxl.Workbook()
            
            # メインシートの作成（パフォーマンス最適化）
            ws = wb.active
            ws.title = "損害賠償計算書"
            
            # 設定から表示項目を取得
            report_items = self.report_config.excel_report_items
            
            # シート作成（メソッド分割でパフォーマンス向上）
            self._create_calculation_sheet(ws, case_data, results, report_items)
            
            # 追加シート作成（設定により制御）
            if "detailed_calculation_table" in report_items:
                self._create_detail_sheet(wb, case_data, results)
            
            if "charts" in report_items and self.report_config.include_charts_in_excel:
                self._create_chart_sheet(wb, results)
            
            if "reference_materials" in report_items:
                self._create_reference_sheet(wb, case_data)
            
            # 列幅の調整とスタイル適用（バッチ処理で最適化）
            self._apply_formatting_batch(ws)
            
            # 会社ロゴの挿入（設定されている場合）
            if self.report_config.company_logo_path and "logo" in report_items:
                self._insert_company_logo(ws)
            
            # Excelファイルのプロパティ設定
            self._set_excel_properties(wb, case_data)
            
            # ファイル保存
            output_path = self._get_output_path(output_filename)
            wb.save(output_path)
            
            self.logger.info(f"Excel帳票を正常に作成しました: {output_path}")
            return True
            
        except Exception as e:
            self.error_handler.handle_exception(
                FileIOError(f"Excel帳票の作成に失敗しました: {e}",
                            user_message="Excel損害賠償計算書の作成中にエラーが発生しました。",
                            context={"filename": output_filename, "template_type": template_type})
            )
            return False

    def _get_output_path(self, output_filename: str) -> Path:
        """出力パスを取得"""
        output_path = self.output_dir / output_filename
        if not output_filename.endswith('.xlsx'):
            output_path = output_path.with_suffix('.xlsx')
        return output_path

    @monitor_performance("excel_sheet_creation", track_parameters=False)
    def _create_calculation_sheet(self, ws, case_data: CaseData, 
                                results: Dict[str, CalculationResult], 
                                report_items: List[str]):
        """計算シートの作成（最適化版）"""
        row = 1
        
        # ヘッダー作成
        if "header_title" in report_items:
            row = self._create_header(ws, row)
        
        # 基本情報テーブル
        if "basic_info_table" in report_items:
            row = self._create_case_info_section(ws, case_data, row)
        
        # 医療情報
        if "medical_info_table" in report_items:
            row = self._create_medical_info_section(ws, case_data, row)
        
        # 収入情報
        if "income_info_table" in report_items:
            row = self._create_income_info_section(ws, case_data, row)
        
        # 計算結果
        if "calculation_results_table" in report_items:
            row = self._create_calculation_results(ws, results, row)
        
        # 詳細計算
        if "detailed_calculation_table" in report_items:
            row = self._create_detailed_calculations(ws, results, row)
        
        # フッター
        if "disclaimer_footer" in report_items:
            row = self._create_footer(ws, row)

    def _create_header(self, ws, start_row: int) -> int:
        """ヘッダー部分の作成 - 最適化版"""
        row = start_row
        
        # タイトル
        ws[f'A{row}'] = "損害賠償計算書（弁護士基準）"
        ws[f'A{row}'].font = self.fonts['title']
        ws[f'A{row}'].alignment = self.alignments['center']
        ws.merge_cells(f'A{row}:H{row}')
        row += 2
        
        # 作成日時
        ws[f'G{row}'] = f"作成日: {datetime.now().strftime('%Y年%m月%d日')}"
        ws[f'G{row}'].font = self.fonts['small']
        ws[f'G{row}'].alignment = self.alignments['right']
        row += 2
        
        return row

    def _create_case_info_section(self, ws, case_data: CaseData, start_row: int) -> int:
        """事件情報セクションの作成"""
        row = start_row
        
        # セクションヘッダー
        ws[f'A{row}'] = "【事件情報】"
        ws[f'A{row}'].font = self.fonts['subtitle']
        ws[f'A{row}'].fill = self.fills['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        # 情報テーブル
        info_data = [
            ("事件番号", case_data.case_number or "未設定"),
            ("被害者氏名", case_data.victim_name or "未設定"),
            ("年齢", f"{case_data.victim_age}歳" if case_data.victim_age else "未設定"),
            ("性別", case_data.victim_gender or "未設定"),
            ("事故日", case_data.accident_date.strftime('%Y年%m月%d日') if case_data.accident_date else "未設定")
        ]
        
        row = self._create_info_table(ws, info_data, row)
        row += 1
        
        return row

    def _create_info_table(self, ws, data: List[tuple], start_row: int) -> int:
        """情報テーブルの作成（共通メソッド）"""
        row = start_row
        
        for label, value in data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            
            # スタイル適用
            ws[f'A{row}'].font = self.fonts['body']
            ws[f'A{row}'].fill = self.fills['subheader']
            ws[f'A{row}'].border = self.borders['thin_all']
            ws[f'A{row}'].alignment = self.alignments['left']
            
            ws[f'B{row}'].font = self.fonts['body']
            ws[f'B{row}'].border = self.borders['thin_all']
            ws[f'B{row}'].alignment = self.alignments['left']
            
            row += 1
        
        return row

    def _create_medical_info_section(self, ws, case_data: CaseData, start_row: int) -> int:
        """医療情報セクションの作成"""
        row = start_row
        
        ws[f'A{row}'] = "【医療情報】"
        ws[f'A{row}'].font = self.fonts['subtitle']
        ws[f'A{row}'].fill = self.fills['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        medical_data = [
            ("入院期間", f"{case_data.hospitalization_days}日" if case_data.hospitalization_days else "なし"),
            ("通院期間", f"{case_data.outpatient_days}日" if case_data.outpatient_days else "なし"),
            ("後遺障害等級", case_data.disability_grade or "なし"),
            ("治療費", f"¥{case_data.medical_expenses:,}" if case_data.medical_expenses else "未確定")
        ]
        
        row = self._create_info_table(ws, medical_data, row)
        row += 1
        
        return row

    def _create_income_info_section(self, ws, case_data: CaseData, start_row: int) -> int:
        """収入情報セクションの作成"""
        row = start_row
        
        ws[f'A{row}'] = "【収入情報】"
        ws[f'A{row}'].font = self.fonts['subtitle']
        ws[f'A{row}'].fill = self.fills['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        income_data = [
            ("年収", f"¥{case_data.annual_income:,}" if case_data.annual_income else "未設定"),
            ("職業", case_data.occupation or "未設定"),
            ("基礎収入", f"¥{case_data.base_income:,}" if case_data.base_income else "未算出")
        ]
        
        row = self._create_info_table(ws, income_data, row)
        row += 1
        
        return row

    @monitor_performance("excel_calculation_results", track_parameters=False)
    def _create_calculation_results(self, ws, results: Dict[str, CalculationResult], start_row: int) -> int:
        """計算結果セクションの作成（パフォーマンス最適化）"""
        row = start_row
        
        # セクションヘッダー
        ws[f'A{row}'] = "【損害賠償額計算結果】"
        ws[f'A{row}'].font = self.fonts['subtitle']
        ws[f'A{row}'].fill = self.fills['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        # テーブルヘッダー
        headers = ["損害項目", "金額", "計算根拠", "法的根拠"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.fonts['header']
            cell.fill = self.fills['header']
            cell.alignment = self.alignments['center']
            cell.border = self.borders['thin_all']
        row += 1
        
        # 各損害項目（バッチ処理で最適化）
        total_amount = Decimal('0')
        for key, result in results.items():
            if key == 'summary':
                continue
                
            # データ行の作成
            ws.cell(row=row, column=1, value=result.item_name)
            ws.cell(row=row, column=2, value=f"¥{result.amount:,}")
            ws.cell(row=row, column=3, value=result.calculation_details)
            ws.cell(row=row, column=4, value=result.legal_basis)
            
            # スタイル適用（バッチ処理）
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = self.borders['thin_all']
                if col == 1:
                    cell.font = self.fonts['body']
                elif col == 2:
                    cell.font = self.fonts['money']
                    cell.alignment = self.alignments['right']
                    if result.amount > 0:
                        cell.fill = self.fills['amount']
                else:
                    cell.font = self.fonts['small']
                    cell.alignment = self.alignments['left']
            
            total_amount += result.amount
            row += 1
        
        # 合計行
        if 'summary' in results:
            summary = results['summary']
            ws.cell(row=row, column=1, value="【総合計】")
            ws.cell(row=row, column=2, value=f"¥{summary.amount:,}")
            
            # 合計行のスタイル
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = self.fonts['header']
                cell.fill = self.fills['total']
                cell.border = self.borders['medium_all']
                if col == 2:
                    cell.alignment = self.alignments['right']
            row += 1
        
        return row

    def _create_detailed_calculations(self, ws, results: Dict[str, CalculationResult], start_row: int) -> int:
        """詳細計算の作成"""
        row = start_row
        
        ws[f'A{row}'] = "【計算詳細】"
        ws[f'A{row}'].font = self.fonts['subtitle']
        ws[f'A{row}'].fill = self.fills['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        for key, result in results.items():
            if key == 'summary':
                continue
            
            # 項目名
            ws[f'A{row}'] = f"■ {result.item_name}"
            ws[f'A{row}'].font = self.fonts['body']
            row += 1
            
            # 計算詳細
            if result.calculation_details:
                details_lines = result.calculation_details.split('\n')
                for line in details_lines:
                    if line.strip():
                        ws[f'B{row}'] = line.strip()
                        ws[f'B{row}'].font = self.fonts['small']
                        row += 1
            
            row += 1  # 項目間のスペース
        
        return row

    def _create_footer(self, ws, start_row: int) -> int:
        """フッター部分の作成"""
        row = start_row
        
        # 作成者情報
        ws[f'A{row}'] = f"作成者: {self.report_config.default_author}"
        ws[f'A{row}'].font = self.fonts['small']
        row += 1
        
        # 作成日時
        ws[f'A{row}'] = f"作成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
        ws[f'A{row}'].font = self.fonts['small']
        row += 1
        
        # 免責事項
        disclaimer = "※ この計算書は弁護士基準に基づく概算であり、実際の示談や裁判では個別事情により変動する可能性があります。"
        ws[f'A{row}'] = disclaimer
        ws[f'A{row}'].font = self.fonts['small']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(f'A{row}:H{row}')
        ws.row_dimensions[row].height = 30
        row += 1
        
        return row

    @monitor_performance("excel_formatting", track_parameters=False)
    def _apply_formatting_batch(self, ws):
        """フォーマット設定の一括適用（パフォーマンス最適化）"""
        try:
            # 列幅設定
            for col_letter, width in self.column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # 行高設定（最初の数行）
            for row_num in range(1, min(6, ws.max_row + 1)):
                ws.row_dimensions[row_num].height = self.row_heights.get('header', 25.0)
                        
        except Exception as e:
            self.logger.warning(f"列幅・行高の設定適用に失敗しました: {e}")

    def _insert_company_logo(self, ws):
        """会社ロゴの挿入"""
        try:
            if self.report_config.company_logo_path:
                logo_path = Path(self.report_config.company_logo_path)
                if logo_path.exists():
                    img = Image(str(logo_path))
                    img.height = 60
                    img.width = 120
                    ws.add_image(img, 'H1')
                    self.logger.debug("会社ロゴを挿入しました")
        except Exception as e:
            self.logger.warning(f"会社ロゴの挿入に失敗しました: {e}")

    def _set_excel_properties(self, wb: openpyxl.Workbook, case_data: CaseData):
        """Excelファイルのプロパティ設定"""
        try:
            props = wb.properties
            props.title = f"損害賠償計算書 - {case_data.case_number or '事件番号未設定'}"
            props.subject = "弁護士基準損害賠償計算"
            props.creator = self.report_config.default_author
            props.description = f"被害者: {case_data.victim_name or '氏名未設定'}"
            props.keywords = "損害賠償,弁護士基準,計算書"
            props.lastModifiedBy = self.report_config.default_author
            props.created = datetime.now()
            props.modified = datetime.now()
            
            self.logger.debug("Excelファイルプロパティを設定しました")
        except Exception as e:
            self.logger.warning(f"Excelファイルプロパティの設定に失敗しました: {e}")

    @monitor_performance("excel_additional_sheets", track_parameters=False)
    def _create_detail_sheet(self, wb: openpyxl.Workbook, case_data: CaseData, 
                           results: Dict[str, CalculationResult]):
        """詳細シートの作成"""
        ws = wb.create_sheet("計算詳細")
        row = 1
        
        for key, result in results.items():
            if key == 'summary':
                continue
            
            # 項目タイトル
            ws.cell(row=row, column=1, value=f"【{result.item_name}】")
            ws.cell(row=row, column=1).font = self.fonts['subtitle']
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # 計算詳細
            if result.calculation_details:
                details_lines = result.calculation_details.split('\n')
                for line in details_lines:
                    if line.strip():
                        ws.cell(row=row, column=1, value=line)
                        ws.cell(row=row, column=1).font = self.fonts['body']
                        row += 1
            
            # 法的根拠
            if result.legal_basis:
                ws.cell(row=row, column=1, value=f"法的根拠: {result.legal_basis}")
                ws.cell(row=row, column=1).font = self.fonts['small']
                row += 1
            
            row += 2

    def _create_chart_sheet(self, wb: openpyxl.Workbook, results: Dict[str, CalculationResult]):
        """グラフシートの作成"""
        ws = wb.create_sheet("グラフ")
        
        # データ準備
        row = 2
        ws.cell(row=1, column=1, value="損害項目")
        ws.cell(row=1, column=2, value="金額")
        
        chart_data = []
        for key, result in results.items():
            if key == 'summary' or result.amount == 0:
                continue
            
            ws.cell(row=row, column=1, value=result.item_name)
            ws.cell(row=row, column=2, value=float(result.amount))
            chart_data.append((result.item_name, float(result.amount)))
            row += 1
        
        # 棒グラフ作成
        if chart_data:
            chart = BarChart()
            chart.title = "損害項目別金額"
            chart.x_axis.title = "損害項目"
            chart.y_axis.title = "金額（円）"
            
            data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
            categories = Reference(ws, min_col=1, min_row=2, max_row=row-1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            ws.add_chart(chart, "D2")

    def _create_reference_sheet(self, wb: openpyxl.Workbook, case_data: CaseData):
        """付属資料シートの作成"""
        ws = wb.create_sheet("付属資料")
        
        explanations = [
            "【弁護士基準（裁判基準）について】",
            "",
            "弁護士基準とは、過去の裁判例に基づいて定められた損害賠償の算定基準です。",
            "一般的に自賠責基準や任意保険基準よりも高額になる傾向があります。",
            "",
            "【計算に使用した基準】",
            "・入通院慰謝料: 赤い本2023年版 別表I/II",
            "・後遺障害慰謝料: 赤い本2023年版",
            "・ライプニッツ係数: 法定利率3%（令和2年4月1日以降）",
            "",
            "【注意事項】",
            "・本計算書は弁護士基準に基づく概算額です",
            "・実際の示談交渉や裁判では個別事情により変動する可能性があります",
            "・最新の法改正や判例により基準が変更される場合があります"
        ]
        
        for i, text in enumerate(explanations, 1):
            ws.cell(row=i, column=1, value=text)
            if text.startswith("【") and text.endswith("】"):
                ws.cell(row=i, column=1).font = self.fonts['header']
            else:
                ws.cell(row=i, column=1).font = self.fonts['body']


class TemplateManager:
    """Excelテンプレート管理クラス（最適化版）"""
    
    def __init__(self, report_config: ReportConfig, error_handler: ErrorHandler):
        self.report_config = report_config
        self.error_handler = error_handler
        self.logger = logging.getLogger(__name__)
        
        # テンプレートディレクトリの設定
        self._setup_template_directory()

    def _setup_template_directory(self):
        """テンプレートディレクトリの設定"""
        try:
            if self.report_config.excel_templates:
                default_template_path = self.report_config.excel_templates.get("default")
                if default_template_path:
                    self.template_dir = Path(default_template_path).parent
                else:
                    self.template_dir = Path("templates/excel")
            else:
                if self.report_config.excel_template_path:
                    self.template_dir = Path(self.report_config.excel_template_path).parent
                else:
                    self.template_dir = Path("templates/excel")
            
            self.template_dir.mkdir(parents=True, exist_ok=True)
            self.logger.info(f"TemplateManager: テンプレートディレクトリを設定: {self.template_dir}")
        except OSError as e:
            self.error_handler.handle_exception(
                FileIOError(f"テンプレートディレクトリの作成/アクセスに失敗しました: {e}",
                            user_message="テンプレート用ディレクトリの準備中にエラーが発生しました。",
                            context={"path": str(self.template_dir)})
            )
            self.template_dir = Path("./templates")
            self.template_dir.mkdir(parents=True, exist_ok=True)

    @monitor_performance("template_creation", track_parameters=True)
    def create_standard_templates(self):
        """標準テンプレートの作成"""
        try:
            self._create_traffic_accident_template()
            self._create_work_accident_template()
            self._create_medical_malpractice_template()
            self.logger.info("標準テンプレートを作成しました。")
        except Exception as e:
            self.error_handler.handle_exception(
                CompensationSystemError(f"標準テンプレートの作成中にエラーが発生しました: {e}",
                                        category=ErrorCategory.FILE_IO,
                                        severity=ErrorSeverity.MEDIUM,
                                        user_message="標準レポートテンプレートの作成に失敗しました。",
                                        context={"details": str(e)})
            )

    def _create_traffic_accident_template(self):
        """交通事故用テンプレート"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "交通事故損害賠償計算書"
            
            ws['A1'] = "交通事故損害賠償計算書（テンプレート）"
            ws['A1'].font = Font(name='Meiryo UI', size=18, bold=True)
            ws.merge_cells('A1:I1')
            
            ws['A2'] = "[ロゴプレースホルダー]"
            ws['H10'] = f"作成者: {self.report_config.default_author}"
            
            template_path = self.template_dir / "traffic_accident_template.xlsx"
            wb.save(template_path)
            self.logger.debug(f"交通事故テンプレートを作成/更新しました: {template_path}")
        except Exception as e:
            self.error_handler.handle_exception(
                FileIOError(f"交通事故テンプレートの作成に失敗しました: {e}",
                            user_message="交通事故用レポートテンプレートの作成に失敗しました。",
                            context={"template_name": "traffic_accident_template.xlsx", "details": str(e)})
            )

    def _create_work_accident_template(self):
        """労災事故用テンプレート"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "労災事故損害賠償計算書"
            
            ws['A1'] = "労災事故損害賠償計算書（テンプレート）"
            ws['A1'].font = Font(name='Meiryo UI', size=18, bold=True)
            ws.merge_cells('A1:I1')
            
            # 労災特有項目の設定
            headers = [
                "事業場情報", "災害発生年月日", "災害の種類", "災害の原因",
                "被災者情報", "療養給付", "休業給付", "障害給付"
            ]
            
            row = 4
            for i, header in enumerate(headers):
                ws.cell(row=row + i, column=1, value=f"{header}:")
                ws.cell(row=row + i, column=2, value="[データプレースホルダー]")
            
            template_path = self.template_dir / "work_accident_template.xlsx"
            wb.save(template_path)
            self.logger.debug(f"労災事故テンプレートを作成/更新しました: {template_path}")
        except Exception as e:
            self.error_handler.handle_exception(
                FileIOError(f"労災事故テンプレートの作成に失敗しました: {e}",
                            user_message="労災事故用レポートテンプレートの作成に失敗しました。",
                            context={"template_name": "work_accident_template.xlsx", "details": str(e)})
            )

    def _create_medical_malpractice_template(self):
        """医療過誤用テンプレート"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "医療過誤損害賠償計算書"
            
            ws['A1'] = "医療過誤損害賠償計算書（テンプレート）"
            ws['A1'].font = Font(name='Meiryo UI', size=18, bold=True)
            ws.merge_cells('A1:I1')
            
            # 医療過誤特有項目の設定
            headers = [
                "医療機関情報", "診療科", "主治医", "診療期間",
                "医療行為の内容", "過誤の内容", "因果関係", "過失の程度"
            ]
            
            row = 4
            for i, header in enumerate(headers):
                ws.cell(row=row + i, column=1, value=f"{header}:")
                ws.cell(row=row + i, column=2, value="[データプレースホルダー]")
            
            template_path = self.template_dir / "medical_malpractice_template.xlsx"
            wb.save(template_path)
            self.logger.debug(f"医療過誤テンプレートを作成/更新しました: {template_path}")
        except Exception as e:
            self.error_handler.handle_exception(
                FileIOError(f"医療過誤テンプレートの作成に失敗しました: {e}",
                            user_message="医療過誤用レポートテンプレートの作成に失敗しました。",
                            context={"template_name": "medical_malpractice_template.xlsx", "details": str(e)})
            )

    @monitor_performance("template_application", track_parameters=True)
    def apply_template(self, template_name: str, case_data: CaseData) -> Optional[openpyxl.Workbook]:
        """テンプレートの適用"""
        template_mapping = {
            "traffic_accident": "traffic_accident_template.xlsx",
            "work_accident": "work_accident_template.xlsx", 
            "medical_malpractice": "medical_malpractice_template.xlsx",
            "default": "traffic_accident_template.xlsx"
        }
        
        template_filename = template_mapping.get(template_name, template_mapping["default"])
        template_path = self.template_dir / template_filename
        
        try:
            if template_path.exists():
                wb = openpyxl.load_workbook(template_path)
                self._populate_template_data(wb, case_data, template_name)
                self.logger.info(f"テンプレート '{template_name}' を適用しました。")
                return wb
            else:
                self.logger.warning(f"テンプレート {template_filename} が見つかりません。作成を試行します。")
                self._create_template_if_missing(template_name)
                
                if template_path.exists():
                    wb = openpyxl.load_workbook(template_path)
                    self._populate_template_data(wb, case_data, template_name)
                    return wb
                else:
                    self.error_handler.handle_exception(
                        FileNotFoundError(f"テンプレート {template_name} が見つからず、作成にも失敗しました: {template_path}"),
                        user_message=f"指定されたレポートテンプレート '{template_name}' が見つかりませんでした。",
                        context={"template_name": template_name, "path": str(template_path)}
                    )
                    return None
        except Exception as e:
            self.error_handler.handle_exception(
                FileIOError(f"テンプレート '{template_name}' の適用中にエラーが発生しました: {e}",
                            user_message=f"レポートテンプレート '{template_name}' の処理中にエラーが発生しました。",
                            context={"template_name": template_name, "path": str(template_path), "details": str(e)})
            )
            return None

    def _create_template_if_missing(self, template_name: str):
        """不足テンプレートの作成"""
        try:
            if template_name == "traffic_accident":
                self._create_traffic_accident_template()
            elif template_name == "work_accident":
                self._create_work_accident_template()
            elif template_name == "medical_malpractice":
                self._create_medical_malpractice_template()
        except Exception as e:
            self.logger.error(f"テンプレート {template_name} の作成に失敗しました: {e}")

    def _populate_template_data(self, wb: openpyxl.Workbook, case_data: CaseData, template_type: str):
        """テンプレートにデータを埋め込む"""
        try:
            ws = wb.active
            
            # 基本情報の埋め込み
            if case_data.case_number:
                self._replace_placeholder(ws, "[事件番号]", case_data.case_number)
            if case_data.victim_name:
                self._replace_placeholder(ws, "[被害者名]", case_data.victim_name)
            if case_data.accident_date:
                self._replace_placeholder(ws, "[事故日]", case_data.accident_date.strftime("%Y年%m月%d日"))
            
            # 作成日
            self._replace_placeholder(ws, "[作成日プレースホルダー]", datetime.now().strftime("%Y年%m月%d日"))
            
        except Exception as e:
            self.logger.error(f"テンプレートデータの埋め込み中にエラーが発生しました: {e}")

    def _replace_placeholder(self, ws: openpyxl.worksheet.worksheet.Worksheet, placeholder: str, value: str):
        """プレースホルダーの置換"""
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and placeholder in cell.value:
                    cell.value = cell.value.replace(placeholder, value)


# パフォーマンス改善された使用例
if __name__ == "__main__":
    # パフォーマンス監視付きのサンプル使用
    generator = ExcelReportGeneratorOptimized()
    
    # テンプレートマネージャー
    template_mgr = TemplateManager(generator.report_config, generator.error_handler)
    template_mgr.create_standard_templates()
    
    # パフォーマンス監視レポートの出力
    performance_monitor = get_performance_monitor()
    
    # サンプル実行後のレポート生成
    import time
    time.sleep(1)  # メトリクス収集のため
    
    summary = performance_monitor.get_performance_summary(hours=1)
    print("=== パフォーマンス要約 ===")
    print(f"関数呼び出し総数: {summary['total_function_calls']}")
    print(f"平均実行時間: {summary['performance']['avg_execution_time']:.3f}秒")
    print(f"最大実行時間: {summary['performance']['max_execution_time']:.3f}秒")
    print(f"エラー率: {summary['performance']['error_rate']:.1%}")
